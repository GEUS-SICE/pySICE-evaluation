# -*- coding: utf-8 -*-
"""
Created on Tue Aug 11 13:23:03 2020

@author: bav
"""

import pandas as pd
import numpy as np
import bav_lib as bl
import rasterio as rio
import os
import matplotlib.pyplot as plt

#%%
def load_cook_data():
    print('loading Cook\'s field measurements')
    data_cook = pd.concat((pd.read_csv('data/Cook_data/Hbio_albedo_JB.csv'), 
                           pd.read_csv('data/Cook_data/Lbio_albedo_JB.csv')),
                          axis=1)
    data_cook=data_cook.drop( columns='Unnamed: 0' )
    
    # creating measurements' labels from column names
    col = data_cook.columns.values.astype('U')
    tmp = pd.DataFrame(np.char.split(col,sep="_"))
    
    labels = pd.DataFrame()
    labels[['bio_load', 'day', 'month', 'sitename']] = pd.DataFrame(tmp[0].tolist(), index= tmp.index)
    labels['samplename'] = labels['bio_load'] +'_'+labels['day'] +'_'+labels['month'] + '_17_'++labels['sitename'] 
    WL = pd.read_csv('data/Cook_data/Wavlengths.csv',header=None).values
    lambda_names = ['lambda_' + str(x[0]) for x in WL]
    
    data_cook=data_cook.transpose()
    data_cook.columns = lambda_names
    data_cook['samplename'] = labels['samplename'].values
    # data_cook['bio_load'] = labels['bio_load'].values
    # data_cook['day'] = labels['day'].values
    # data_cook['month'] = labels['month'].values
    # data_cook['sitename'] = labels['sitename'].values
    data_cook=data_cook.reset_index()
    data_cook=data_cook.set_index(['samplename'])
    data_cook=data_cook.drop(columns='index')
    
    
    print('Loading Cook\'s metadata') 
    from openpyxl import load_workbook
    wb = load_workbook(filename='data/Cook_data/metadata_JB.xlsx', 
                       read_only=True)
    ws = wb['Sheet1']
    
    # Read the cell values into a list of lists
    data_rows = []
    for row in ws['A2':'H18']:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)
    
    # Transform into dataframe
    df = pd.DataFrame(data_rows)
    df.columns=df.iloc[0,:].values
    df=df.drop(index=0,axis=1)
    # df[['day','month','year',"sitename"]]= df["Site Name"].str.split("_", n = 4, expand = True) 
    df['bio_load']= 'HA'
    
    data_rows = []
    for row in ws['A21':'H34']:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)
    
    # Transform into dataframe
    df2 = pd.DataFrame(data_rows)
    df2.columns=df2.iloc[0,:].values
    df2=df2.drop(index=0,axis=1)
    # df2[['day','month','year',"sitename"]]= df2["Site Name"].str.split("_", n = 4, expand = True) 
    df2['bio_load']= 'LA'
    
    metadata = pd.concat((df,df2))
    metadata['samplename']= metadata['bio_load']+'_'+metadata['Site Name']
    metadata=metadata.set_index(['samplename'])
    metadata=metadata.drop(columns=['Site Name','bio_load'])
    metadata= metadata.drop('HA_14_7_17_SB7',axis=0)
    metadata= metadata.loc[metadata['Cloud Cover'].str.extract('(\d+)').astype(int).values<=3]
    return(data_cook, metadata, WL)

#%% 
def plot_sice_output(InputFolder, var_list = ('albedo_bb_planar_sw','albedo_bb_spherical_sw')):
   #% Plotting output
    try:
        os.mkdir(InputFolder+'plots')
    except:
        print('folder exist')
    fig,ax=bl.heatmap_discrete(rio.open(InputFolder+'diagnostic_retrieval.tif').read(1),
                            'diagnostic_retrieval ')
    ax.set_title(InputFolder)
    fig.savefig(InputFolder+'plots/diagnostic_retrieval.png',bbox_inches='tight')
    
    
    for i in range(len(var_list)):
        var_1 = rio.open(InputFolder+var_list[i]+'.tif').read(1)
        plt.figure(figsize=(10,15))
        bl.heatmap(var_1,var_list[i], col_lim=(0, 1) ,cmap_in='jet')
        plt.title(InputFolder)
        plt.savefig(InputFolder+'plots/'+var_list[i]+'.png',bbox_inches='tight')
        plt.close()
        
    var_list = ('O3_SICE', 'grain_diameter', 'snow_specific_area',
                'al','conc','r0')
    for i in range(len(var_list)):
        var_1 = rio.open(InputFolder+var_list[i]+'.tif').read(1)
        plt.figure(figsize=(10,15))
        bl.heatmap(var_1,var_list[i],cmap_in='jet')
        plt.title(InputFolder)
        plt.savefig(InputFolder+'plots/'+var_list[i]+'.png',bbox_inches='tight')
        plt.close()
    
    for i in np.arange(21):
        var_name = 'albedo_spectral_spherical_'+ str(i+1).zfill(2)
        try:
            var_1 = rio.open(InputFolder+var_name+'.tif').read(1)
        except: 
            continue
        plt.figure(figsize=(10,15))
        bl.heatmap(var_1,var_name, col_lim=(0, 1) ,cmap_in='jet')
        plt.title(InputFolder)
        plt.savefig(InputFolder+'plots/'+var_name+'.png',bbox_inches='tight')
        plt.close()
        var_name = 'albedo_spectral_planar_'+ str(i+1).zfill(2)
        var_1 = rio.open(InputFolder+var_name+'.tif').read(1)
        plt.figure(figsize=(10,15))
        bl.heatmap(var_1,var_name, col_lim=(0, 1) ,cmap_in='jet')
        plt.title(InputFolder)
        plt.savefig(InputFolder+'plots/'+var_name+'.png',bbox_inches='tight')
        plt.close()
        var_name = 'rBRR_'+ str(i+1).zfill(2)
        var_1 = rio.open(InputFolder+var_name+'.tif').read(1)
        plt.figure(figsize=(10,15))
        bl.heatmap(var_1,var_name, col_lim=(0, 1) ,cmap_in='jet')
        plt.title(InputFolder)
        plt.savefig(InputFolder+'plots/'+var_name+'.png',bbox_inches='tight')
        plt.close()
        var_name = 'r_TOA_'+ str(i+1).zfill(2)
        var_1 = rio.open(InputFolder+var_name+'.tif').read(1)
        plt.figure(figsize=(10,15))
        bl.heatmap(var_1,var_name, col_lim=(0, 1) ,cmap_in='jet')
        plt.title(InputFolder)
        plt.savefig(InputFolder+'plots/'+var_name+'.png',bbox_inches='tight')
        plt.close()